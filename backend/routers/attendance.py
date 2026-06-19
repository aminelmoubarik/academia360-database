from datetime import datetime, date, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from mysql.connector import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from auth import require_roles
from db import get_db
from models import (
    AttendanceCreate,
    AttendancePunchRequest,
    AttendanceUpdate,
    OfflineAttendanceSyncRequest,
)
from utils import get_audit_username, model_to_dict

router = APIRouter(prefix="/attendance", tags=["Attendance"])


def validate_student_exists(cursor, student_id: int):
    cursor.execute("""
        SELECT StudentID
        FROM Tbl_Students
        WHERE StudentID = %s
    """, (student_id,))

    if cursor.fetchone() is None:
        raise HTTPException(status_code=404, detail="Student not found")


def validate_schedule_exists(cursor, schedule_id: int | None):
    if schedule_id is None:
        return

    cursor.execute("""
        SELECT ScheduleID
        FROM Tbl_GeneratedSchedule
        WHERE ScheduleID = %s
    """, (schedule_id,))

    if cursor.fetchone() is None:
        raise HTTPException(status_code=404, detail="Schedule record not found")


def get_student_by_card(cursor, card_uid: str):
    cleaned = card_uid.strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="Card UID is required")

    cursor.execute("""
        SELECT
            s.StudentID AS student_id,
            s.FullName AS student_name,
            s.StudentNumber AS student_number,
            s.CardUID AS card_uid,
            s.ClassID AS class_id,
            cl.Name AS class_name,
            c.Code AS course_code
        FROM Tbl_Students s
        LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
        LEFT JOIN Tbl_Courses c ON cl.CourseID = c.CourseID
        WHERE s.CardUID = %s OR s.StudentNumber = %s
        LIMIT 1
    """, (cleaned, cleaned))

    student = cursor.fetchone()
    if student is None:
        raise HTTPException(
            status_code=404,
            detail="No student found for this card/identifier"
        )
    return student


def resolve_auto_punch_type(cursor, student_id: int, punch_time: datetime | None):
    target = punch_time or datetime.now()
    cursor.execute("""
        SELECT PunchType
        FROM Tbl_AttendanceRecords
        WHERE StudentID = %s
          AND DATE(PunchTime) = %s
        ORDER BY PunchTime DESC, AttendanceRecordID DESC
        LIMIT 1
    """, (student_id, target.date()))

    last = cursor.fetchone()
    if last is not None and last["PunchType"] == "in":
        return "out"
    return "in"


def find_current_schedule(cursor, class_id: int | None, punch_time: datetime | None):
    if class_id is None:
        return None

    target = punch_time or datetime.now()
    target_date = target.date()
    target_time = target.time()

    cursor.execute("""
        SELECT gs.ScheduleID AS schedule_id
        FROM Tbl_GeneratedSchedule gs
        JOIN Tbl_SchoolCalendar sc ON gs.CalendarID = sc.CalendarID
        WHERE gs.ClassID = %s
          AND sc.CalendarDate = %s
          AND gs.StartTime <= %s
          AND gs.EndTime >= %s
          AND gs.Status <> 'cancelled'
        ORDER BY gs.StartTime
        LIMIT 1
    """, (class_id, target_date, target_time, target_time))

    schedule = cursor.fetchone()
    return None if schedule is None else schedule["schedule_id"]


def insert_attendance_record(
    cursor,
    *,
    student_id: int,
    schedule_id: int | None,
    punch_type: str,
    punch_method: str,
    punch_time: datetime | None,
    is_synced: bool,
    audit_username: str,
):
    cursor.execute("""
        INSERT INTO Tbl_AttendanceRecords (
            StudentID,
            ScheduleID,
            PunchType,
            PunchMethod,
            PunchTime,
            IsSynced,
            InsertUsername
        )
        VALUES (%s, %s, %s, %s, COALESCE(%s, CURRENT_TIMESTAMP), %s, %s)
    """, (
        student_id,
        schedule_id,
        punch_type,
        punch_method,
        punch_time,
        is_synced,
        audit_username,
    ))
    return cursor.lastrowid


def build_punch_response(cursor, attendance_record_id: int, student: dict, punch_type: str, schedule_id: int | None):
    cursor.execute("""
        SELECT PunchTime AS punch_time
        FROM Tbl_AttendanceRecords
        WHERE AttendanceRecordID = %s
    """, (attendance_record_id,))
    punch = cursor.fetchone() or {}
    return {
        "success": True,
        "message": "Attendance punch registered successfully",
        "attendance_record_id": attendance_record_id,
        "student": student,
        "punch_type": punch_type,
        "schedule_id": schedule_id,
        "punch_time": punch.get("punch_time"),
    }


def _format_export_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    value_text = str(value)
    if len(value_text) >= 8 and value_text.count(":") >= 2:
        return value_text[:5]
    return value_text


def _attendance_type_label(value: str | None) -> str:
    return {"in": "Entrada", "out": "Saída"}.get(value or "", value or "")


def _attendance_method_label(value: str | None) -> str:
    return {
        "nfc": "NFC",
        "rfid": "RFID",
        "qr": "QR Code",
        "barcode": "Código de barras",
        "manual": "Manual",
    }.get(value or "", value or "")


def _safe_filename_part(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.lower())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "assiduidade"


def _attendance_export_filename(rows, extension: str) -> str:
    today = datetime.now().strftime("%Y%m%d")
    class_name = "assiduidade"
    if rows:
        first_class = rows[0].get("class_name")
        if first_class:
            class_name = first_class
    return f"academia360_{_safe_filename_part(class_name)}_{today}.{extension}"


def _validate_attendance_filters(punch_type: str | None, punch_method: str | None):
    allowed_types = {"in", "out"}
    allowed_methods = {"nfc", "rfid", "qr", "barcode", "manual"}

    if punch_type is not None and punch_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid punch type")
    if punch_method is not None and punch_method not in allowed_methods:
        raise HTTPException(status_code=400, detail="Invalid punch method")


def fetch_attendance_rows(
    cursor,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    class_id: int | None = None,
    student_id: int | None = None,
    discipline_id: int | None = None,
    punch_type: str | None = None,
    punch_method: str | None = None,
    is_synced: bool | None = None,
    search: str | None = None,
    limit: int | None = 300,
):
    _validate_attendance_filters(punch_type, punch_method)

    where_clauses = []
    values = []

    if start_date is not None:
        where_clauses.append("DATE(ar.PunchTime) >= %s")
        values.append(start_date)
    if end_date is not None:
        where_clauses.append("DATE(ar.PunchTime) <= %s")
        values.append(end_date)
    if class_id is not None:
        where_clauses.append("cl.ClassID = %s")
        values.append(class_id)
    if student_id is not None:
        where_clauses.append("s.StudentID = %s")
        values.append(student_id)
    if discipline_id is not None:
        where_clauses.append("dc.DisciplineID = %s")
        values.append(discipline_id)
    if punch_type is not None:
        where_clauses.append("ar.PunchType = %s")
        values.append(punch_type)
    if punch_method is not None:
        where_clauses.append("ar.PunchMethod = %s")
        values.append(punch_method)
    if is_synced is not None:
        where_clauses.append("ar.IsSynced = %s")
        values.append(1 if is_synced else 0)
    if search is not None and search.strip():
        like = f"%{search.strip().lower()}%"
        where_clauses.append("""
            (
                LOWER(s.FullName) LIKE %s
                OR LOWER(s.StudentNumber) LIKE %s
                OR LOWER(COALESCE(s.CardUID, '')) LIKE %s
                OR LOWER(COALESCE(cl.Name, '')) LIKE %s
                OR LOWER(COALESCE(c.Code, '')) LIKE %s
                OR LOWER(COALESCE(d.Name, '')) LIKE %s
                OR LOWER(ar.PunchMethod) LIKE %s
                OR LOWER(ar.PunchType) LIKE %s
            )
        """)
        values.extend([like] * 8)

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    limit_sql = ""
    if limit is not None:
        limit_sql = "LIMIT %s"
        values.append(limit)

    cursor.execute(f"""
        SELECT
            ar.AttendanceRecordID AS id,
            ar.StudentID AS student_id,
            s.FullName AS student_name,
            s.StudentNumber AS student_number,
            s.CardUID AS card_uid,
            cl.ClassID AS class_id,
            cl.Name AS class_name,
            c.Code AS course_code,
            c.Name AS course_name,
            sy.Name AS school_year,
            ar.ScheduleID AS schedule_id,
            dc.DisciplineID AS discipline_id,
            d.Name AS discipline_name,
            sc.CalendarDate AS schedule_date,
            gs.StartTime AS class_start_time,
            gs.EndTime AS class_end_time,
            ar.PunchType AS punch_type,
            ar.PunchMethod AS punch_method,
            ar.PunchTime AS punch_time,
            ar.IsSynced AS is_synced,
            ar.InsertUsername AS insert_username,
            ar.InsertDate AS insert_date,
            ar.ChangeUsername AS change_username,
            ar.ChangeDate AS change_date
        FROM Tbl_AttendanceRecords ar
        JOIN Tbl_Students s
            ON ar.StudentID = s.StudentID
        LEFT JOIN Tbl_Classes cl
            ON s.ClassID = cl.ClassID
        LEFT JOIN Tbl_Courses c
            ON cl.CourseID = c.CourseID
        LEFT JOIN Tref_SchoolYears sy
            ON cl.SchoolYearID = sy.SchoolYearID
        LEFT JOIN Tbl_GeneratedSchedule gs
            ON ar.ScheduleID = gs.ScheduleID
        LEFT JOIN Tbl_SchoolCalendar sc
            ON gs.CalendarID = sc.CalendarID
        LEFT JOIN trx_Discipline_CourseYear dc
            ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
        LEFT JOIN Tbl_Disciplines d
            ON dc.DisciplineID = d.DisciplineID
        {where_sql}
        ORDER BY ar.PunchTime DESC
        {limit_sql}
    """, tuple(values))

    return cursor.fetchall()


@router.get("")
def get_attendance_records(
    start_date: date | None = None,
    end_date: date | None = None,
    class_id: int | None = None,
    student_id: int | None = None,
    discipline_id: int | None = None,
    punch_type: str | None = None,
    punch_method: str | None = None,
    is_synced: bool | None = None,
    search: str | None = None,
    limit: int = 300,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    limit = max(1, min(limit, 1000))
    cursor = connection.cursor(dictionary=True)

    try:
        return fetch_attendance_rows(
            cursor,
            start_date=start_date,
            end_date=end_date,
            class_id=class_id,
            student_id=student_id,
            discipline_id=discipline_id,
            punch_type=punch_type,
            punch_method=punch_method,
            is_synced=is_synced,
            search=search,
            limit=limit,
        )

    finally:
        cursor.close()


@router.get("/dashboard")
def get_attendance_dashboard(
    target_date: date | None = None,
    class_id: int | None = None,
    discipline_id: int | None = None,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    target_date = target_date or date.today()
    cursor = connection.cursor(dictionary=True)

    try:
        student_where = []
        student_values = []
        if class_id is not None:
            student_where.append("s.ClassID = %s")
            student_values.append(class_id)
        student_where_sql = ""
        if student_where:
            student_where_sql = "WHERE " + " AND ".join(student_where)

        cursor.execute(f"""
            SELECT COUNT(*) AS total_students
            FROM Tbl_Students s
            {student_where_sql}
        """, tuple(student_values))
        total_students = (cursor.fetchone() or {}).get("total_students", 0)

        attendance_where = ["DATE(ar.PunchTime) = %s", "ar.PunchType = 'in'"]
        attendance_values = [target_date]
        if class_id is not None:
            attendance_where.append("s.ClassID = %s")
            attendance_values.append(class_id)
        if discipline_id is not None:
            attendance_where.append("dc.DisciplineID = %s")
            attendance_values.append(discipline_id)
        attendance_where_sql = " AND ".join(attendance_where)

        cursor.execute(f"""
            SELECT COUNT(DISTINCT ar.StudentID) AS present_students
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s ON ar.StudentID = s.StudentID
            LEFT JOIN Tbl_GeneratedSchedule gs ON ar.ScheduleID = gs.ScheduleID
            LEFT JOIN trx_Discipline_CourseYear dc ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            WHERE {attendance_where_sql}
        """, tuple(attendance_values))
        present_students = (cursor.fetchone() or {}).get("present_students", 0)

        not_exists_filters = ["ar2.StudentID = s.StudentID", "DATE(ar2.PunchTime) = %s", "ar2.PunchType = 'in'"]
        not_exists_values = [target_date]
        if discipline_id is not None:
            not_exists_filters.append("dc2.DisciplineID = %s")
            not_exists_values.append(discipline_id)

        cursor.execute(f"""
            SELECT
                s.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                s.CardUID AS card_uid,
                cl.Name AS class_name
            FROM Tbl_Students s
            LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
            WHERE NOT EXISTS (
                SELECT 1
                FROM Tbl_AttendanceRecords ar2
                LEFT JOIN Tbl_GeneratedSchedule gs2 ON ar2.ScheduleID = gs2.ScheduleID
                LEFT JOIN trx_Discipline_CourseYear dc2 ON gs2.DisciplineCourseYearID = dc2.DisciplineCourseYearID
                WHERE {' AND '.join(not_exists_filters)}
            )
            {('AND s.ClassID = %s') if class_id is not None else ''}
            ORDER BY cl.Name, s.FullName
            LIMIT 20
        """, tuple([*not_exists_values, *student_values]))
        absentees = cursor.fetchall()

        recent_where = ["DATE(ar.PunchTime) = %s"]
        recent_values = [target_date]
        if class_id is not None:
            recent_where.append("s.ClassID = %s")
            recent_values.append(class_id)
        if discipline_id is not None:
            recent_where.append("dc.DisciplineID = %s")
            recent_values.append(discipline_id)

        cursor.execute(f"""
            SELECT
                ar.AttendanceRecordID AS id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                cl.Name AS class_name,
                d.Name AS discipline_name,
                ar.PunchType AS punch_type,
                ar.PunchMethod AS punch_method,
                ar.PunchTime AS punch_time
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s ON ar.StudentID = s.StudentID
            LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
            LEFT JOIN Tbl_GeneratedSchedule gs ON ar.ScheduleID = gs.ScheduleID
            LEFT JOIN trx_Discipline_CourseYear dc ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            LEFT JOIN Tbl_Disciplines d ON dc.DisciplineID = d.DisciplineID
            WHERE {' AND '.join(recent_where)}
            ORDER BY ar.PunchTime DESC
            LIMIT 12
        """, tuple(recent_values))
        recent = cursor.fetchall()

        cursor.execute("""
            SELECT
                cl.ClassID AS class_id,
                cl.Name AS class_name,
                COUNT(DISTINCT s.StudentID) AS total_students,
                COUNT(DISTINCT CASE
                    WHEN DATE(ar.PunchTime) = %s AND ar.PunchType = 'in' THEN ar.StudentID
                END) AS present_students
            FROM Tbl_Classes cl
            LEFT JOIN Tbl_Students s ON s.ClassID = cl.ClassID
            LEFT JOIN Tbl_AttendanceRecords ar ON ar.StudentID = s.StudentID
            GROUP BY cl.ClassID, cl.Name
            ORDER BY cl.Name
        """, (target_date,))
        by_class = cursor.fetchall()

        cursor.execute("""
            SELECT
                d.DisciplineID AS discipline_id,
                d.Name AS discipline_name,
                COUNT(ar.AttendanceRecordID) AS punch_count,
                COUNT(DISTINCT ar.StudentID) AS student_count
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_GeneratedSchedule gs ON ar.ScheduleID = gs.ScheduleID
            JOIN trx_Discipline_CourseYear dc ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Students s ON ar.StudentID = s.StudentID
            WHERE DATE(ar.PunchTime) = %s
              AND (%s IS NULL OR s.ClassID = %s)
            GROUP BY d.DisciplineID, d.Name
            ORDER BY punch_count DESC, d.Name
            LIMIT 8
        """, (target_date, class_id, class_id))
        by_discipline = cursor.fetchall()

        cursor.execute("""
            SELECT
                ar.PunchMethod AS punch_method,
                COUNT(*) AS total
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s ON ar.StudentID = s.StudentID
            WHERE DATE(ar.PunchTime) = %s
              AND (%s IS NULL OR s.ClassID = %s)
            GROUP BY ar.PunchMethod
            ORDER BY total DESC
        """, (target_date, class_id, class_id))
        by_method = cursor.fetchall()

        absent_students = max(total_students - present_students, 0)
        return {
            "date": target_date,
            "filters": {
                "class_id": class_id,
                "discipline_id": discipline_id,
            },
            "total_students": total_students,
            "present_students": present_students,
            "absent_students": absent_students,
            "absentees": absentees,
            "recent": recent,
            "by_class": by_class,
            "by_discipline": by_discipline,
            "by_method": by_method,
            "alerts": [
                {
                    "type": "absenteeism",
                    "message": f"{absent_students} estudantes sem picagem de entrada no dia selecionado",
                }
            ] if absent_students > 0 else [],
        }

    finally:
        cursor.close()


@router.get("/alerts")
def get_absenteeism_alerts(
    target_date: date | None = None,
    class_id: int | None = None,
    discipline_id: int | None = None,
    days: int = 7,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    """Return actionable absenteeism alerts for the selected date and recent period."""
    target_date = target_date or date.today()
    days = max(1, min(days, 30))
    period_start = target_date - timedelta(days=days - 1)
    cursor = connection.cursor(dictionary=True)

    try:
        student_filters = []
        student_values = []
        if class_id is not None:
            student_filters.append("s.ClassID = %s")
            student_values.append(class_id)
        student_where = ""
        if student_filters:
            student_where = "AND " + " AND ".join(student_filters)

        discipline_join = ""
        discipline_filter = ""
        discipline_values = []
        if discipline_id is not None:
            discipline_join = """
                LEFT JOIN Tbl_GeneratedSchedule gs2 ON ar2.ScheduleID = gs2.ScheduleID
                LEFT JOIN trx_Discipline_CourseYear dc2 ON gs2.DisciplineCourseYearID = dc2.DisciplineCourseYearID
            """
            discipline_filter = "AND dc2.DisciplineID = %s"
            discipline_values.append(discipline_id)

        cursor.execute(f"""
            SELECT
                s.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                s.CardUID AS card_uid,
                cl.Name AS class_name
            FROM Tbl_Students s
            LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
            WHERE NOT EXISTS (
                SELECT 1
                FROM Tbl_AttendanceRecords ar2
                {discipline_join}
                WHERE ar2.StudentID = s.StudentID
                  AND DATE(ar2.PunchTime) = %s
                  AND ar2.PunchType = 'in'
                  {discipline_filter}
            )
            {student_where}
            ORDER BY cl.Name, s.FullName
            LIMIT 50
        """, tuple([target_date, *discipline_values, *student_values]))
        today_absentees = cursor.fetchall()

        cursor.execute(f"""
            SELECT
                s.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                cl.Name AS class_name,
                COUNT(sc.CalendarDate) AS missing_days
            FROM Tbl_Students s
            LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
            JOIN Tbl_SchoolCalendar sc
                ON sc.CalendarDate BETWEEN %s AND %s
               AND sc.IsSchoolDay = TRUE
            LEFT JOIN Tbl_AttendanceRecords ar
                ON ar.StudentID = s.StudentID
               AND DATE(ar.PunchTime) = sc.CalendarDate
               AND ar.PunchType = 'in'
            WHERE ar.AttendanceRecordID IS NULL
            {student_where}
            GROUP BY s.StudentID, s.FullName, s.StudentNumber, cl.Name
            HAVING missing_days >= 2
            ORDER BY missing_days DESC, cl.Name, s.FullName
            LIMIT 30
        """, tuple([period_start, target_date, *student_values]))
        recurrent_absences = cursor.fetchall()

        cursor.execute(f"""
            SELECT
                cl.ClassID AS class_id,
                cl.Name AS class_name,
                COUNT(DISTINCT s.StudentID) AS total_students,
                COUNT(DISTINCT CASE WHEN ar.AttendanceRecordID IS NULL THEN s.StudentID END) AS missing_today
            FROM Tbl_Students s
            LEFT JOIN Tbl_Classes cl ON s.ClassID = cl.ClassID
            LEFT JOIN Tbl_AttendanceRecords ar
                ON ar.StudentID = s.StudentID
               AND DATE(ar.PunchTime) = %s
               AND ar.PunchType = 'in'
            WHERE 1 = 1
            {student_where}
            GROUP BY cl.ClassID, cl.Name
            HAVING missing_today > 0
            ORDER BY missing_today DESC, cl.Name
            LIMIT 10
        """, tuple([target_date, *student_values]))
        class_alerts = cursor.fetchall()

        alert_items = []
        if today_absentees:
            alert_items.append({
                "type": "daily_absence",
                "severity": "high" if len(today_absentees) >= 5 else "medium",
                "message": f"{len(today_absentees)} estudantes sem entrada registada no dia selecionado",
            })
        if recurrent_absences:
            alert_items.append({
                "type": "recurrent_absence",
                "severity": "high",
                "message": f"{len(recurrent_absences)} estudantes com ausências recorrentes nos últimos {days} dias",
            })

        return {
            "date": target_date,
            "period_start": period_start,
            "period_days": days,
            "filters": {
                "class_id": class_id,
                "discipline_id": discipline_id,
            },
            "today_absentees": today_absentees,
            "recurrent_absences": recurrent_absences,
            "class_alerts": class_alerts,
            "alerts": alert_items,
        }

    finally:
        cursor.close()


@router.post("/punch")
def punch_attendance(
    request: AttendancePunchRequest,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        student = get_student_by_card(cursor, request.card_uid)
        punch_type = request.punch_type or resolve_auto_punch_type(
            cursor=cursor,
            student_id=student["student_id"],
            punch_time=request.punch_time,
        )
        schedule_id = find_current_schedule(
            cursor=cursor,
            class_id=student.get("class_id"),
            punch_time=request.punch_time,
        )
        attendance_id = insert_attendance_record(
            cursor,
            student_id=student["student_id"],
            schedule_id=schedule_id,
            punch_type=punch_type,
            punch_method=request.punch_method,
            punch_time=request.punch_time,
            is_synced=request.is_synced,
            audit_username=audit_username,
        )
        connection.commit()
        return build_punch_response(cursor, attendance_id, student, punch_type, schedule_id)

    except HTTPException:
        connection.rollback()
        raise
    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))
    finally:
        cursor.close()


@router.post("/offline-sync")
def sync_offline_attendance(
    request: OfflineAttendanceSyncRequest,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)
    synced = []
    failed = []

    try:
        for index, item in enumerate(request.records):
            try:
                student = get_student_by_card(cursor, item.card_uid)
                punch_type = item.punch_type or resolve_auto_punch_type(
                    cursor,
                    student["student_id"],
                    item.punch_time,
                )
                schedule_id = find_current_schedule(cursor, student.get("class_id"), item.punch_time)
                attendance_id = insert_attendance_record(
                    cursor,
                    student_id=student["student_id"],
                    schedule_id=schedule_id,
                    punch_type=punch_type,
                    punch_method=item.punch_method,
                    punch_time=item.punch_time,
                    is_synced=True,
                    audit_username=audit_username,
                )
                synced.append({"index": index, "attendance_record_id": attendance_id})
            except Exception as exc:  # noqa: BLE001 - queremos devolver os erros por registo
                failed.append({"index": index, "card_uid": item.card_uid, "error": str(exc)})

        # Commit successful records even when some records fail.
        # This makes offline sync resilient: valid punches are not lost because
        # one card UID is invalid or one record has a bad timestamp.
        connection.commit()

        if failed:
            return {
                "success": False,
                "synced": synced,
                "failed": failed,
                "message": "Partial sync completed. Some records need attention.",
            }

        return {"success": True, "synced": synced, "failed": []}

    finally:
        cursor.close()


@router.get("/export/excel")
def export_attendance_excel(
    start_date: Optional[date] = Query(default=None),
    end_date: Optional[date] = Query(default=None),
    class_id: Optional[int] = Query(default=None),
    student_id: Optional[int] = Query(default=None),
    discipline_id: Optional[int] = Query(default=None),
    punch_type: Optional[str] = Query(default=None),
    punch_method: Optional[str] = Query(default=None),
    is_synced: Optional[bool] = Query(default=None),
    search: Optional[str] = Query(default=None),
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        rows = fetch_attendance_rows(
            cursor,
            start_date=start_date,
            end_date=end_date,
            class_id=class_id,
            student_id=student_id,
            discipline_id=discipline_id,
            punch_type=punch_type,
            punch_method=punch_method,
            is_synced=is_synced,
            search=search,
            limit=5000,
        )

        if not rows:
            raise HTTPException(
                status_code=404,
                detail="Não foram encontrados registos de assiduidade para exportar."
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Assiduidade"

        sheet.merge_cells("A1:L1")
        title_cell = sheet["A1"]
        title_cell.value = "Academia360 - Relatório de Assiduidade"
        title_cell.font = Font(bold=True, size=16, color="1C29E1")
        title_cell.alignment = Alignment(horizontal="center")

        generated_cell = sheet["A2"]
        generated_cell.value = f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        generated_cell.font = Font(italic=True, color="666666")

        filter_cell = sheet["A3"]
        filter_cell.value = (
            f"Período: {start_date or 'início'} até {end_date or 'fim'} · "
            f"Turma: {class_id or 'todas'} · Disciplina: {discipline_id or 'todas'}"
        )
        filter_cell.font = Font(italic=True, color="666666")

        headers = [
            "Data/hora",
            "Estudante",
            "N.º estudante",
            "Cartão",
            "Turma",
            "Curso",
            "Ano letivo",
            "Disciplina",
            "Tipo",
            "Método",
            "Sincronização",
            "Criado por",
        ]
        sheet.append(headers)
        header_row = 4

        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1C29E1")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            sheet.append([
                _format_export_value(row.get("punch_time")),
                row.get("student_name"),
                row.get("student_number"),
                row.get("card_uid"),
                row.get("class_name"),
                row.get("course_code"),
                row.get("school_year"),
                row.get("discipline_name"),
                _attendance_type_label(row.get("punch_type")),
                _attendance_method_label(row.get("punch_method")),
                "Sim" if row.get("is_synced") else "Offline",
                row.get("insert_username"),
            ])

        widths = [20, 30, 16, 18, 16, 14, 16, 26, 12, 18, 16, 18]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:L{sheet.max_row}"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = _attendance_export_filename(rows, "xlsx")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    finally:
        cursor.close()


@router.get("/export/pdf")
def export_attendance_pdf(
    start_date: Optional[date] = Query(default=None),
    end_date: Optional[date] = Query(default=None),
    class_id: Optional[int] = Query(default=None),
    student_id: Optional[int] = Query(default=None),
    discipline_id: Optional[int] = Query(default=None),
    punch_type: Optional[str] = Query(default=None),
    punch_method: Optional[str] = Query(default=None),
    is_synced: Optional[bool] = Query(default=None),
    search: Optional[str] = Query(default=None),
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        rows = fetch_attendance_rows(
            cursor,
            start_date=start_date,
            end_date=end_date,
            class_id=class_id,
            student_id=student_id,
            discipline_id=discipline_id,
            punch_type=punch_type,
            punch_method=punch_method,
            is_synced=is_synced,
            search=search,
            limit=5000,
        )

        if not rows:
            raise HTTPException(
                status_code=404,
                detail="Não foram encontrados registos de assiduidade para exportar."
            )

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.0 * cm,
            leftMargin=1.0 * cm,
            topMargin=1.0 * cm,
            bottomMargin=1.0 * cm,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Academia360 - Relatório de Assiduidade", styles["Title"]),
            Paragraph(f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
            Paragraph(
                f"Período: {start_date or 'início'} até {end_date or 'fim'} · "
                f"Turma: {class_id or 'todas'} · Disciplina: {discipline_id or 'todas'}",
                styles["Normal"],
            ),
            Spacer(1, 0.35 * cm),
        ]

        data = [["Data/hora", "Estudante", "Turma", "Disciplina", "Tipo", "Método", "Sync"]]
        for row in rows:
            data.append([
                _format_export_value(row.get("punch_time")),
                row.get("student_name") or "",
                row.get("class_name") or "",
                row.get("discipline_name") or "",
                _attendance_type_label(row.get("punch_type")),
                _attendance_method_label(row.get("punch_method")),
                "Sim" if row.get("is_synced") else "Offline",
            ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[3.0*cm, 5.2*cm, 3.0*cm, 4.4*cm, 2.0*cm, 2.6*cm, 1.8*cm],
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C29E1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDE2FF")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FC")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))

        elements.append(table)
        document.build(elements)
        buffer.seek(0)

        filename = _attendance_export_filename(rows, "pdf")
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    finally:
        cursor.close()


@router.get("/student/{student_id}")
def get_attendance_by_student(
    student_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        validate_student_exists(cursor, student_id)

        cursor.execute("""
            SELECT
                ar.AttendanceRecordID AS id,
                ar.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                ar.ScheduleID AS schedule_id,
                d.Name AS discipline_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS class_start_time,
                gs.EndTime AS class_end_time,
                ar.PunchType AS punch_type,
                ar.PunchMethod AS punch_method,
                ar.PunchTime AS punch_time,
                ar.IsSynced AS is_synced
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s 
                ON ar.StudentID = s.StudentID
            LEFT JOIN Tbl_GeneratedSchedule gs 
                ON ar.ScheduleID = gs.ScheduleID
            LEFT JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            LEFT JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            LEFT JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            WHERE ar.StudentID = %s
            ORDER BY ar.PunchTime DESC
        """, (student_id,))

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/schedule/{schedule_id}")
def get_attendance_by_schedule(
    schedule_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        validate_schedule_exists(cursor, schedule_id)

        cursor.execute("""
            SELECT
                ar.AttendanceRecordID AS id,
                ar.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                ar.ScheduleID AS schedule_id,
                d.Name AS discipline_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS class_start_time,
                gs.EndTime AS class_end_time,
                ar.PunchType AS punch_type,
                ar.PunchMethod AS punch_method,
                ar.PunchTime AS punch_time,
                ar.IsSynced AS is_synced
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s 
                ON ar.StudentID = s.StudentID
            LEFT JOIN Tbl_GeneratedSchedule gs 
                ON ar.ScheduleID = gs.ScheduleID
            LEFT JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            LEFT JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            LEFT JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            WHERE ar.ScheduleID = %s
            ORDER BY ar.PunchTime DESC
        """, (schedule_id,))

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/{attendance_id}")
def get_attendance_record(
    attendance_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                ar.AttendanceRecordID AS id,
                ar.StudentID AS student_id,
                s.FullName AS student_name,
                s.StudentNumber AS student_number,
                s.CardUID AS card_uid,
                cl.ClassID AS class_id,
                cl.Name AS class_name,
                c.Code AS course_code,
                sy.Name AS school_year,
                ar.ScheduleID AS schedule_id,
                d.Name AS discipline_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS class_start_time,
                gs.EndTime AS class_end_time,
                ar.PunchType AS punch_type,
                ar.PunchMethod AS punch_method,
                ar.PunchTime AS punch_time,
                ar.IsSynced AS is_synced,
                ar.InsertUsername AS insert_username,
                ar.InsertDate AS insert_date,
                ar.ChangeUsername AS change_username,
                ar.ChangeDate AS change_date
            FROM Tbl_AttendanceRecords ar
            JOIN Tbl_Students s 
                ON ar.StudentID = s.StudentID
            LEFT JOIN Tbl_Classes cl 
                ON s.ClassID = cl.ClassID
            LEFT JOIN Tbl_Courses c 
                ON cl.CourseID = c.CourseID
            LEFT JOIN Tref_SchoolYears sy 
                ON cl.SchoolYearID = sy.SchoolYearID
            LEFT JOIN Tbl_GeneratedSchedule gs 
                ON ar.ScheduleID = gs.ScheduleID
            LEFT JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            LEFT JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            LEFT JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            WHERE ar.AttendanceRecordID = %s
        """, (attendance_id,))

        record = cursor.fetchone()

        if record is None:
            raise HTTPException(
                status_code=404,
                detail="Attendance record not found"
            )

        return record

    finally:
        cursor.close()


@router.post("")
def create_attendance_record(
    attendance: AttendanceCreate,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        validate_student_exists(cursor, attendance.student_id)
        validate_schedule_exists(cursor, attendance.schedule_id)

        attendance_id = insert_attendance_record(
            cursor,
            student_id=attendance.student_id,
            schedule_id=attendance.schedule_id,
            punch_type=attendance.punch_type,
            punch_method=attendance.punch_method,
            punch_time=attendance.punch_time,
            is_synced=attendance.is_synced,
            audit_username=audit_username,
        )

        connection.commit()

        return {
            "message": "Attendance record created successfully",
            "attendance_record_id": attendance_id
        }

    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))

    finally:
        cursor.close()


@router.put("/{attendance_id}")
def update_attendance_record(
    attendance_id: int,
    attendance: AttendanceUpdate,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "secretary"]))
):
    data = model_to_dict(attendance)

    if not data:
        raise HTTPException(status_code=400, detail="No fields provided for update")

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        cursor.execute("""
            SELECT AttendanceRecordID
            FROM Tbl_AttendanceRecords
            WHERE AttendanceRecordID = %s
        """, (attendance_id,))

        if cursor.fetchone() is None:
            raise HTTPException(
                status_code=404,
                detail="Attendance record not found"
            )

        if "student_id" in data:
            validate_student_exists(cursor, data["student_id"])

        if "schedule_id" in data:
            validate_schedule_exists(cursor, data["schedule_id"])

        field_map = {
            "student_id": "StudentID",
            "schedule_id": "ScheduleID",
            "punch_type": "PunchType",
            "punch_method": "PunchMethod",
            "is_synced": "IsSynced"
        }

        set_clauses = []
        values = []

        for api_field, db_field in field_map.items():
            if api_field in data:
                set_clauses.append(f"{db_field} = %s")
                values.append(data[api_field])

        if not set_clauses:
            raise HTTPException(status_code=400, detail="No valid fields provided for update")

        set_clauses.append("ChangeUsername = %s")
        values.append(audit_username)

        values.append(attendance_id)

        cursor.execute(f"""
            UPDATE Tbl_AttendanceRecords
            SET {", ".join(set_clauses)}
            WHERE AttendanceRecordID = %s
        """, tuple(values))

        connection.commit()
  
        return {
            "message": "Attendance record updated successfully",
            "attendance_record_id": attendance_id
        }

    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))

    finally:
        cursor.close()


@router.delete("/{attendance_id}")
def delete_attendance_record(
    attendance_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "secretary"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            DELETE FROM Tbl_AttendanceRecords
            WHERE AttendanceRecordID = %s
        """, (attendance_id,))

        connection.commit()

        if cursor.rowcount == 0:
            raise HTTPException(
                status_code=404,
                detail="Attendance record not found"
            )

        return {
            "message": "Attendance record deleted successfully",
            "attendance_record_id": attendance_id
        }

    except IntegrityError:
        connection.rollback()
        raise HTTPException(
            status_code=400,
            detail="Attendance record cannot be deleted because it is being used by another record"
        )

    finally:
        cursor.close()
