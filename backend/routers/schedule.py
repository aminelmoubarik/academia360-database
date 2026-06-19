from datetime import date, datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from mysql.connector import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from auth import require_roles
from db import get_db
from models import ScheduleApprovalRequest, ScheduleCreate, ScheduleGenerateRequest, ScheduleUpdate
from services.schedule_generator import generate_schedule_algorithm
from utils import get_audit_username, model_to_dict

router = APIRouter(prefix="/schedule", tags=["Schedule"])


def validate_time_range(start_time, end_time):
    if start_time is not None and end_time is not None:
        if end_time <= start_time:
            raise HTTPException(
                status_code=400,
                detail="A hora de fim deve ser superior à hora de início"
            )


def get_existing_schedule(cursor, schedule_id: int):
    cursor.execute("""
        SELECT
            ScheduleID,
            ClassID,
            DisciplineCourseYearID,
            ProfessorID,
            RoomID,
            CalendarID,
            StartTime,
            EndTime,
            Status
        FROM Tbl_GeneratedSchedule
        WHERE ScheduleID = %s
    """, (schedule_id,))

    schedule = cursor.fetchone()

    if schedule is None:
        raise HTTPException(status_code=404, detail="Registo de horário não encontrado")

    return schedule


def validate_schedule_dependencies(
    cursor,
    class_id,
    discipline_course_year_id,
    professor_id,
    room_id,
    calendar_id
):
    cursor.execute("""
        SELECT
            ClassID,
            CourseID,
            SchoolYearID,
            CourseYearNumber
        FROM Tbl_Classes
        WHERE ClassID = %s
    """, (class_id,))
    class_record = cursor.fetchone()

    if class_record is None:
        raise HTTPException(status_code=404, detail="Turma não encontrada")

    cursor.execute("""
        SELECT
            DisciplineCourseYearID,
            CourseID,
            SchoolYearID,
            CourseYearNumber,
            IsPractical
        FROM trx_Discipline_CourseYear
        WHERE DisciplineCourseYearID = %s
    """, (discipline_course_year_id,))
    discipline_course_year = cursor.fetchone()

    if discipline_course_year is None:
        raise HTTPException(
            status_code=404,
            detail="Configuração da disciplina/curso/ano não encontrada"
        )

    cursor.execute("""
        SELECT ProfessorID
        FROM Tbl_Professors
        WHERE ProfessorID = %s
    """, (professor_id,))
    professor = cursor.fetchone()

    if professor is None:
        raise HTTPException(status_code=404, detail="Professor não encontrado")

    cursor.execute("""
        SELECT
            RoomID,
            IsPracticeRoom
        FROM Tbl_Rooms
        WHERE RoomID = %s
    """, (room_id,))
    room = cursor.fetchone()

    if room is None:
        raise HTTPException(status_code=404, detail="Sala não encontrada")

    cursor.execute("""
        SELECT
            CalendarID,
            SchoolYearID,
            CalendarDate,
            IsSchoolDay
        FROM Tbl_SchoolCalendar
        WHERE CalendarID = %s
    """, (calendar_id,))
    calendar = cursor.fetchone()

    if calendar is None:
        raise HTTPException(status_code=404, detail="Registo de calendário não encontrado")

    if not calendar["IsSchoolDay"]:
        raise HTTPException(
            status_code=400,
            detail="Não é possível criar horário num dia não letivo"
        )

    if class_record["CourseID"] != discipline_course_year["CourseID"]:
        raise HTTPException(
            status_code=400,
            detail="O curso da turma não corresponde ao curso da disciplina"
        )

    if class_record["SchoolYearID"] != discipline_course_year["SchoolYearID"]:
        raise HTTPException(
            status_code=400,
            detail="O ano letivo da turma não corresponde ao ano letivo da disciplina"
        )

    if class_record["CourseYearNumber"] != discipline_course_year["CourseYearNumber"]:
        raise HTTPException(
            status_code=400,
            detail="O ano do curso da turma não corresponde ao ano do curso da disciplina"
        )

    if calendar["SchoolYearID"] != class_record["SchoolYearID"]:
        raise HTTPException(
            status_code=400,
            detail="O ano letivo do calendário não corresponde ao ano letivo da turma"
        )

    cursor.execute("""
        SELECT 1
        FROM trx_Professor_DisciplineCourseYear
        WHERE ProfessorID = %s
          AND DisciplineCourseYearID = %s
    """, (
        professor_id,
        discipline_course_year_id
    ))
    professor_assignment = cursor.fetchone()

    if professor_assignment is None:
        raise HTTPException(
            status_code=400,
            detail="O professor não está atribuído a esta disciplina/curso/ano"
        )

    if discipline_course_year["IsPractical"] and not room["IsPracticeRoom"]:
        raise HTTPException(
            status_code=400,
            detail="Disciplinas práticas devem ser marcadas numa sala prática"
        )


def validate_schedule_conflicts(
    cursor,
    class_id,
    professor_id,
    room_id,
    calendar_id,
    start_time,
    end_time,
    excluded_schedule_id=None
):
    excluded_clause = ""
    excluded_values = []

    if excluded_schedule_id is not None:
        excluded_clause = "AND ScheduleID <> %s"
        excluded_values = [excluded_schedule_id]

    cursor.execute(f"""
        SELECT ScheduleID
        FROM Tbl_GeneratedSchedule
        WHERE ClassID = %s
          AND CalendarID = %s
          AND NOT (EndTime <= %s OR StartTime >= %s)
          {excluded_clause}
        LIMIT 1
    """, (
        class_id,
        calendar_id,
        start_time,
        end_time,
        *excluded_values
    ))

    if cursor.fetchone() is not None:
        raise HTTPException(
            status_code=409,
            detail="Conflito de horário: esta turma já tem outra aula neste horário"
        )

    cursor.execute(f"""
        SELECT ScheduleID
        FROM Tbl_GeneratedSchedule
        WHERE ProfessorID = %s
          AND CalendarID = %s
          AND NOT (EndTime <= %s OR StartTime >= %s)
          {excluded_clause}
        LIMIT 1
    """, (
        professor_id,
        calendar_id,
        start_time,
        end_time,
        *excluded_values
    ))

    if cursor.fetchone() is not None:
        raise HTTPException(
            status_code=409,
            detail="Conflito de horário: este professor já tem outra aula neste horário"
        )

    cursor.execute(f"""
        SELECT ScheduleID
        FROM Tbl_GeneratedSchedule
        WHERE RoomID = %s
          AND CalendarID = %s
          AND NOT (EndTime <= %s OR StartTime >= %s)
          {excluded_clause}
        LIMIT 1
    """, (
        room_id,
        calendar_id,
        start_time,
        end_time,
        *excluded_values
    ))

    if cursor.fetchone() is not None:
        raise HTTPException(
            status_code=409,
            detail="Conflito de horário: esta sala já está ocupada neste horário"
        )


def get_class_for_generation(cursor, class_id: int):
    cursor.execute("""
        SELECT
            cl.ClassID AS class_id,
            cl.CourseID AS course_id,
            cl.SchoolYearID AS school_year_id,
            cl.CourseYearNumber AS course_year_number,
            COUNT(s.StudentID) AS class_size
        FROM Tbl_Classes cl
        LEFT JOIN Tbl_Students s
            ON cl.ClassID = s.ClassID
        WHERE cl.ClassID = %s
        GROUP BY
            cl.ClassID,
            cl.CourseID,
            cl.SchoolYearID,
            cl.CourseYearNumber
    """, (class_id,))

    class_data = cursor.fetchone()

    if class_data is None:
        raise HTTPException(status_code=404, detail="Turma não encontrada")

    if class_data["class_size"] == 0:
        class_data["class_size"] = 1

    return class_data


def get_disciplines_for_generation(cursor, class_data):
    cursor.execute("""
        SELECT
            dc.DisciplineCourseYearID AS discipline_course_year_id,
            d.Name AS discipline_name,
            dc.TotalMinutes AS total_minutes,
            dc.LessonDurationMinutes AS lesson_duration_minutes,
            dc.IsPractical AS is_practical
        FROM trx_Discipline_CourseYear dc
        JOIN Tbl_Disciplines d
            ON dc.DisciplineID = d.DisciplineID
        WHERE dc.CourseID = %s
          AND dc.SchoolYearID = %s
          AND dc.CourseYearNumber = %s
        ORDER BY d.Name
    """, (
        class_data["course_id"],
        class_data["school_year_id"],
        class_data["course_year_number"]
    ))

    disciplines = cursor.fetchall()

    if not disciplines:
        raise HTTPException(
            status_code=400,
            detail="Não foram encontradas disciplinas para esta turma/ano do curso"
        )

    for discipline in disciplines:
        discipline["is_practical"] = bool(discipline["is_practical"])

    return disciplines


def get_school_days_for_generation(cursor, school_year_id: int, start_date, end_date):
    cursor.execute("""
        SELECT
            CalendarID AS calendar_id,
            CalendarDate AS date
        FROM Tbl_SchoolCalendar
        WHERE SchoolYearID = %s
          AND CalendarDate BETWEEN %s AND %s
          AND IsSchoolDay = 1
        ORDER BY CalendarDate
    """, (
        school_year_id,
        start_date,
        end_date
    ))

    school_days = cursor.fetchall()

    if not school_days:
        raise HTTPException(
            status_code=400,
            detail="Não foram encontrados dias letivos válidos no intervalo selecionado"
        )

    return school_days


def get_professors_for_generation(cursor, discipline_course_year_ids):
    if not discipline_course_year_ids:
        return []

    placeholders = ", ".join(["%s"] * len(discipline_course_year_ids))

    cursor.execute(f"""
        SELECT
            ProfessorID AS professor_id,
            DisciplineCourseYearID AS discipline_course_year_id
        FROM trx_Professor_DisciplineCourseYear
        WHERE DisciplineCourseYearID IN ({placeholders})
    """, tuple(discipline_course_year_ids))

    professors = cursor.fetchall()

    if not professors:
        raise HTTPException(
            status_code=400,
            detail="Não há professores atribuídos às disciplinas selecionadas"
        )

    return professors


def get_rooms_for_generation(cursor):
    cursor.execute("""
        SELECT
            RoomID AS room_id,
            COALESCE(Capacity, 999) AS capacity,
            IsPracticeRoom AS is_practical
        FROM Tbl_Rooms
        ORDER BY RoomID
    """)

    rooms = cursor.fetchall()

    if not rooms:
        raise HTTPException(
            status_code=400,
            detail="Não há salas disponíveis"
        )

    for room in rooms:
        room["is_practical"] = bool(room["is_practical"])

    return rooms


def get_teacher_availability_for_generation(cursor, school_year_id: int):
    cursor.execute("""
        SELECT
            ProfessorID AS professor_id,
            DayOfWeek AS day_of_week,
            StartTime AS start_time,
            EndTime AS end_time
        FROM Tbl_TeacherAvailability
        WHERE SchoolYearID = %s
    """, (school_year_id,))

    rows = cursor.fetchall()

    day_map = {
        "monday": 0,
        "tuesday": 1,
        "wednesday": 2,
        "thursday": 3,
        "friday": 4
    }

    availability = []

    for row in rows:
        availability.append({
            "professor_id": row["professor_id"],
            "weekday": day_map[row["day_of_week"]],
            "start_time": row["start_time"],
            "end_time": row["end_time"]
        })

    return availability


def get_existing_schedule_for_generation(cursor, start_date, end_date, excluded_class_id=None):
    excluded_clause = ""
    values = [start_date, end_date]

    if excluded_class_id is not None:
        excluded_clause = "AND gs.ClassID <> %s"
        values.append(excluded_class_id)

    cursor.execute(f"""
        SELECT
            gs.ClassID AS class_id,
            gs.ProfessorID AS professor_id,
            gs.RoomID AS room_id,
            gs.CalendarID AS calendar_id,
            sc.CalendarDate AS date,
            gs.StartTime AS start_time,
            gs.EndTime AS end_time
        FROM Tbl_GeneratedSchedule gs
        JOIN Tbl_SchoolCalendar sc
            ON gs.CalendarID = sc.CalendarID
        WHERE sc.CalendarDate BETWEEN %s AND %s
          {excluded_clause}
    """, tuple(values))

    return cursor.fetchall()


def has_existing_schedule_for_class(cursor, class_id: int, start_date, end_date):
    cursor.execute("""
        SELECT 1
        FROM Tbl_GeneratedSchedule gs
        JOIN Tbl_SchoolCalendar sc
            ON gs.CalendarID = sc.CalendarID
        WHERE gs.ClassID = %s
          AND sc.CalendarDate BETWEEN %s AND %s
        LIMIT 1
    """, (
        class_id,
        start_date,
        end_date
    ))

    return cursor.fetchone() is not None


def delete_existing_schedule_for_class(cursor, class_id: int, start_date, end_date):
    cursor.execute("""
        DELETE gs
        FROM Tbl_GeneratedSchedule gs
        JOIN Tbl_SchoolCalendar sc
            ON gs.CalendarID = sc.CalendarID
        WHERE gs.ClassID = %s
          AND sc.CalendarDate BETWEEN %s AND %s
    """, (
        class_id,
        start_date,
        end_date
    ))



def _format_export_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    value_text = str(value)
    if len(value_text) >= 8 and value_text.count(":") >= 2:
        return value_text[:5]
    return value_text


def _status_label(status: str) -> str:
    labels = {
        "draft": "Rascunho",
        "approved": "Aprovado",
        "cancelled": "Rejeitado",
    }
    return labels.get(status or "", status or "")


def _safe_filename_part(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.lower())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "horario"


def fetch_schedule_export_rows(
    cursor,
    class_id: Optional[int] = None,
    professor_id: Optional[int] = None,
    room_id: Optional[int] = None,
    status: str = "approved",
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    allowed_statuses = {"all", "draft", "approved", "cancelled"}
    if status not in allowed_statuses:
        raise HTTPException(
            status_code=400,
            detail="Estado inválido. Use all, draft, approved ou cancelled."
        )

    clauses = []
    values = []

    if class_id is not None:
        clauses.append("gs.ClassID = %s")
        values.append(class_id)
    if professor_id is not None:
        clauses.append("gs.ProfessorID = %s")
        values.append(professor_id)
    if room_id is not None:
        clauses.append("gs.RoomID = %s")
        values.append(room_id)
    if status != "all":
        clauses.append("gs.Status = %s")
        values.append(status)
    if start_date is not None:
        clauses.append("sc.CalendarDate >= %s")
        values.append(start_date)
    if end_date is not None:
        clauses.append("sc.CalendarDate <= %s")
        values.append(end_date)

    where_sql = ""
    if clauses:
        where_sql = "WHERE " + " AND ".join(clauses)

    cursor.execute(f"""
        SELECT
            cl.Name AS class_name,
            c.Name AS course_name,
            sy.Name AS school_year,
            sc.CalendarDate AS schedule_date,
            gs.StartTime AS start_time,
            gs.EndTime AS end_time,
            d.Name AS discipline_name,
            u.FullName AS professor_name,
            r.Name AS room_name,
            gs.Status AS status
        FROM Tbl_GeneratedSchedule gs
        JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
        JOIN Tbl_Courses c ON cl.CourseID = c.CourseID
        JOIN Tref_SchoolYears sy ON cl.SchoolYearID = sy.SchoolYearID
        JOIN trx_Discipline_CourseYear dc
            ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
        JOIN Tbl_Disciplines d
            ON dc.DisciplineID = d.DisciplineID
        JOIN Tbl_Professors p
            ON gs.ProfessorID = p.ProfessorID
        JOIN Tbl_Users u
            ON p.UserID = u.UserID
        JOIN Tbl_Rooms r
            ON gs.RoomID = r.RoomID
        JOIN Tbl_SchoolCalendar sc
            ON gs.CalendarID = sc.CalendarID
        {where_sql}
        ORDER BY cl.Name, sc.CalendarDate, gs.StartTime
    """, tuple(values))

    rows = cursor.fetchall()
    if not rows:
        raise HTTPException(
            status_code=404,
            detail="Não foram encontrados horários para exportar com os filtros selecionados."
        )

    return rows


def _schedule_export_filename(rows, extension: str) -> str:
    class_name = rows[0].get("class_name") or "horario"
    today = datetime.now().strftime("%Y%m%d")
    return f"academia360_{_safe_filename_part(class_name)}_{today}.{extension}"



@router.get("/generation-readiness/{class_id}")
def get_generation_readiness(
    class_id: int,
    start_date: date,
    end_date: date,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    if end_date < start_date:
        raise HTTPException(
            status_code=400,
            detail="A data final deve ser igual ou posterior à data inicial"
        )

    cursor = connection.cursor(dictionary=True)

    try:
        checks = []

        class_data = get_class_for_generation(cursor=cursor, class_id=class_id)
        checks.append({
            "name": "Turma",
            "ok": True,
            "detail": f"Turma {class_id} encontrada com {class_data['class_size']} estudante(s)"
        })

        try:
            disciplines = get_disciplines_for_generation(cursor=cursor, class_data=class_data)
        except HTTPException as error:
            disciplines = []
            checks.append({
                "name": "Disciplinas",
                "ok": False,
                "detail": str(error.detail)
            })
        else:
            checks.append({
                "name": "Disciplinas",
                "ok": bool(disciplines),
                "detail": f"{len(disciplines)} registo(s) de carga horária encontrados"
            })

        try:
            school_days = get_school_days_for_generation(
                cursor=cursor,
                school_year_id=class_data["school_year_id"],
                start_date=start_date,
                end_date=end_date
            )
        except HTTPException as error:
            school_days = []
            checks.append({
                "name": "Calendário escolar",
                "ok": False,
                "detail": str(error.detail)
            })
        else:
            checks.append({
                "name": "Calendário escolar",
                "ok": bool(school_days),
                "detail": f"{len(school_days)} dia(s) letivos válidos no intervalo selecionado"
            })

        discipline_course_year_ids = [
            discipline["discipline_course_year_id"]
            for discipline in disciplines
        ]

        try:
            professors = get_professors_for_generation(
                cursor=cursor,
                discipline_course_year_ids=discipline_course_year_ids
            )
        except HTTPException as error:
            professors = []
            checks.append({
                "name": "Professores atribuídos",
                "ok": False,
                "detail": str(error.detail)
            })
        else:
            professor_conflicts = []
            for discipline in disciplines:
                assigned_count = len([
                    professor for professor in professors
                    if professor["discipline_course_year_id"] == discipline["discipline_course_year_id"]
                ])
                if assigned_count == 0:
                    professor_conflicts.append(discipline["discipline_name"])

            checks.append({
                "name": "Professores atribuídos",
                "ok": not professor_conflicts,
                "detail": (
                    "Todas as disciplinas têm pelo menos um professor"
                    if not professor_conflicts
                    else "Falta atribuir professor a: " + ", ".join(professor_conflicts)
                )
            })

        try:
            rooms = get_rooms_for_generation(cursor)
        except HTTPException as error:
            rooms = []
            checks.append({
                "name": "Salas",
                "ok": False,
                "detail": str(error.detail)
            })
        else:
            room_conflicts = []
            for discipline in disciplines:
                valid_rooms = [
                    room for room in rooms
                    if room["capacity"] >= class_data["class_size"]
                    and (not discipline["is_practical"] or room["is_practical"])
                ]
                if not valid_rooms:
                    room_conflicts.append(discipline["discipline_name"])

            checks.append({
                "name": "Salas",
                "ok": not room_conflicts,
                "detail": (
                    "As salas cumprem os requisitos de capacidade/prática"
                    if not room_conflicts
                    else "Não há sala válida para: " + ", ".join(room_conflicts)
                )
            })

        teacher_availability = get_teacher_availability_for_generation(
            cursor=cursor,
            school_year_id=class_data["school_year_id"]
        )
        checks.append({
            "name": "Disponibilidade",
            "ok": bool(teacher_availability),
            "detail": f"{len(teacher_availability)} janela(s) de disponibilidade encontradas"
        })

        existing = has_existing_schedule_for_class(
            cursor=cursor,
            class_id=class_id,
            start_date=start_date,
            end_date=end_date
        )
        checks.append({
            "name": "Horário existente",
            "ok": not existing,
            "detail": (
                "Não existe horário neste intervalo"
                if not existing
                else "Já existe horário neste intervalo; ative replace_existing para regenerar"
            )
        })

        blocking_checks = [check for check in checks if check["name"] != "Horário existente"]

        return {
            "ready": all(check["ok"] for check in blocking_checks),
            "class_id": class_id,
            "start_date": start_date,
            "end_date": end_date,
            "checks": checks,
            "recommended_test_body": {
                "class_id": class_id,
                "start_date": start_date,
                "end_date": end_date,
                "school_start": "09:00:00",
                "school_end": "17:00:00",
                "replace_existing": True,
                "dry_run": True,
                "status": "draft",
                "max_sessions_per_discipline": 1,
                "max_total_sessions": 300
            }
        }

    finally:
        cursor.close()


@router.post("/generate")
def generate_schedule(
    request: ScheduleGenerateRequest,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin"]))
):
    if request.end_date < request.start_date:
        raise HTTPException(
            status_code=400,
            detail="A data final deve ser igual ou posterior à data inicial"
        )

    validate_time_range(request.school_start, request.school_end)

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        if request.replace_existing and not request.dry_run:
            delete_existing_schedule_for_class(
                cursor=cursor,
                class_id=request.class_id,
                start_date=request.start_date,
                end_date=request.end_date
            )
        elif not request.replace_existing:
            if has_existing_schedule_for_class(
                cursor=cursor,
                class_id=request.class_id,
                start_date=request.start_date,
                end_date=request.end_date
            ):
                raise HTTPException(
                    status_code=409,
                    detail="Já existe horário para esta turma no intervalo selecionado. Use replace_existing=true para regenerar."
                )

        class_data = get_class_for_generation(
            cursor=cursor,
            class_id=request.class_id
        )

        disciplines = get_disciplines_for_generation(
            cursor=cursor,
            class_data=class_data
        )

        school_days = get_school_days_for_generation(
            cursor=cursor,
            school_year_id=class_data["school_year_id"],
            start_date=request.start_date,
            end_date=request.end_date
        )

        discipline_course_year_ids = [
            discipline["discipline_course_year_id"]
            for discipline in disciplines
        ]

        professors = get_professors_for_generation(
            cursor=cursor,
            discipline_course_year_ids=discipline_course_year_ids
        )

        rooms = get_rooms_for_generation(cursor)

        teacher_availability = get_teacher_availability_for_generation(
            cursor=cursor,
            school_year_id=class_data["school_year_id"]
        )

        if not teacher_availability:
            raise HTTPException(
                status_code=400,
                detail="Não existem registos de disponibilidade de professores para este ano letivo"
            )

        existing_schedule = get_existing_schedule_for_generation(
            cursor=cursor,
            start_date=request.start_date,
            end_date=request.end_date,
            excluded_class_id=request.class_id if request.replace_existing else None
        )

        result = generate_schedule_algorithm(
            class_data=class_data,
            disciplines=disciplines,
            school_days=school_days,
            professors=professors,
            rooms=rooms,
            teacher_availability=teacher_availability,
            existing_schedule=existing_schedule,
            school_start=request.school_start,
            school_end=request.school_end,
            max_sessions_per_discipline=request.max_sessions_per_discipline,
            max_total_sessions=request.max_total_sessions
        )

        if not result["success"]:
            connection.rollback()
            return {
                "success": False,
                "message": "A geração do horário falhou",
                "created_records": 0,
                "score": result["score"],
                "conflicts": result["conflicts"],
                "stats": result.get("stats", {}),
                "generation_settings": {
                    "school_start": request.school_start,
                    "school_end": request.school_end,
                    "max_sessions_per_discipline": request.max_sessions_per_discipline,
                    "max_total_sessions": request.max_total_sessions,
                    "dry_run": request.dry_run
                }
            }

        if request.dry_run:
            connection.rollback()
            return {
                "success": True,
                "message": "Horário gerado com sucesso (modo de teste, não guardado)",
                "created_records": result["created_records"],
                "score": result["score"],
                "conflicts": result["conflicts"],
                "stats": result.get("stats", {}),
                "generation_settings": {
                    "school_start": request.school_start,
                    "school_end": request.school_end,
                    "max_sessions_per_discipline": request.max_sessions_per_discipline,
                    "max_total_sessions": request.max_total_sessions,
                    "dry_run": request.dry_run
                },
                "schedule_preview": result["schedule"]
            }

        for record in result["schedule"]:
            cursor.execute("""
                INSERT INTO Tbl_GeneratedSchedule (
                    ClassID,
                    DisciplineCourseYearID,
                    ProfessorID,
                    RoomID,
                    CalendarID,
                    StartTime,
                    EndTime,
                    Status,
                    InsertUsername
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record["class_id"],
                record["discipline_course_year_id"],
                record["professor_id"],
                record["room_id"],
                record["calendar_id"],
                record["start_time"],
                record["end_time"],
                request.status,
                audit_username
            ))

        connection.commit()

        return {
            "success": True,
            "message": "Horário gerado com sucesso",
            "created_records": result["created_records"],
            "score": result["score"],
            "conflicts": result["conflicts"],
            "stats": result.get("stats", {}),
            "generation_settings": {
                "school_start": request.school_start,
                "school_end": request.school_end,
                "max_sessions_per_discipline": request.max_sessions_per_discipline,
                "max_total_sessions": request.max_total_sessions,
                "dry_run": request.dry_run,
                "status": request.status
            }
        }

    except HTTPException:
        connection.rollback()
        raise

    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))

    finally:
        cursor.close()


@router.get("")
def get_schedule(
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.ClassID AS class_id,
                cl.Name AS class_name,
                c.CourseID AS course_id,
                c.Code AS course_code,
                c.Name AS course_name,
                sy.SchoolYearID AS school_year_id,
                sy.Name AS school_year,
                gs.DisciplineCourseYearID AS discipline_course_year_id,
                d.DisciplineID AS discipline_id,
                d.Name AS discipline_name,
                d.Code AS discipline_code,
                gs.ProfessorID AS professor_id,
                u.FullName AS professor_name,
                u.Email AS professor_email,
                gs.RoomID AS room_id,
                r.Name AS room_name,
                gs.CalendarID AS calendar_id,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status,
                gs.InsertUsername AS insert_username,
                gs.InsertDate AS insert_date,
                gs.ChangeUsername AS change_username,
                gs.ChangeDate AS change_date
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN Tbl_Courses c ON cl.CourseID = c.CourseID
            JOIN Tref_SchoolYears sy ON cl.SchoolYearID = sy.SchoolYearID
            JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p 
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u 
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r 
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            ORDER BY sc.CalendarDate, gs.StartTime
        """)

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/class/{class_id}")
def get_schedule_by_class(
    class_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.ClassID AS class_id,
                cl.Name AS class_name,
                d.Name AS discipline_name,
                u.FullName AS professor_name,
                r.Name AS room_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p 
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u 
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r 
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            WHERE gs.ClassID = %s
            ORDER BY sc.CalendarDate, gs.StartTime
        """, (class_id,))

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/professor/{professor_id}")
def get_schedule_by_professor(
    professor_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.ProfessorID AS professor_id,
                u.FullName AS professor_name,
                cl.Name AS class_name,
                d.Name AS discipline_name,
                r.Name AS room_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p 
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u 
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r 
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            WHERE gs.ProfessorID = %s
            ORDER BY sc.CalendarDate, gs.StartTime
        """, (professor_id,))

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/room/{room_id}")
def get_schedule_by_room(
    room_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.RoomID AS room_id,
                r.Name AS room_name,
                cl.Name AS class_name,
                d.Name AS discipline_name,
                u.FullName AS professor_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p 
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u 
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r 
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            WHERE gs.RoomID = %s
            ORDER BY sc.CalendarDate, gs.StartTime
        """, (room_id,))

        return cursor.fetchall()

    finally:
        cursor.close()



@router.get("/pending-approval")
def get_pending_schedule_approvals(
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director"]))
):
    """Return draft timetable groups waiting for Director approval."""
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ClassID AS class_id,
                cl.Name AS class_name,
                c.Name AS course_name,
                sy.Name AS school_year,
                MIN(sc.CalendarDate) AS start_date,
                MAX(sc.CalendarDate) AS end_date,
                COUNT(gs.ScheduleID) AS sessions_count,
                COUNT(DISTINCT gs.DisciplineCourseYearID) AS disciplines_count,
                COUNT(DISTINCT gs.ProfessorID) AS professors_count,
                COUNT(DISTINCT gs.RoomID) AS rooms_count,
                MIN(gs.InsertDate) AS created_at,
                MAX(gs.ChangeDate) AS last_change_at
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN Tbl_Courses c ON cl.CourseID = c.CourseID
            JOIN Tref_SchoolYears sy ON cl.SchoolYearID = sy.SchoolYearID
            JOIN Tbl_SchoolCalendar sc ON gs.CalendarID = sc.CalendarID
            WHERE gs.Status = 'draft'
            GROUP BY
                gs.ClassID,
                cl.Name,
                c.Name,
                sy.Name
            ORDER BY MIN(sc.CalendarDate), cl.Name
        """)

        return cursor.fetchall()

    finally:
        cursor.close()


@router.get("/approval-details/{class_id}")
def get_schedule_approval_details(
    class_id: int,
    start_date: date,
    end_date: date,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.ClassID AS class_id,
                cl.Name AS class_name,
                d.Name AS discipline_name,
                u.FullName AS professor_name,
                r.Name AS room_name,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN trx_Discipline_CourseYear dc
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc
                ON gs.CalendarID = sc.CalendarID
            WHERE gs.ClassID = %s
              AND sc.CalendarDate BETWEEN %s AND %s
              AND gs.Status = 'draft'
            ORDER BY sc.CalendarDate, gs.StartTime
        """, (class_id, start_date, end_date))

        return cursor.fetchall()

    finally:
        cursor.close()


@router.post("/approve-batch")
def approve_schedule_batch(
    request: ScheduleApprovalRequest,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director"]))
):
    if request.end_date < request.start_date:
        raise HTTPException(
            status_code=400,
            detail="A data final deve ser igual ou posterior à data inicial"
        )

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        cursor.execute("""
            UPDATE Tbl_GeneratedSchedule gs
            JOIN Tbl_SchoolCalendar sc
                ON gs.CalendarID = sc.CalendarID
            SET
                gs.Status = 'approved',
                gs.ChangeUsername = %s
            WHERE gs.ClassID = %s
              AND sc.CalendarDate BETWEEN %s AND %s
              AND gs.Status = 'draft'
        """, (
            audit_username,
            request.class_id,
            request.start_date,
            request.end_date
        ))

        connection.commit()

        if cursor.rowcount == 0:
            raise HTTPException(
                status_code=404,
                detail="Não foram encontrados horários em rascunho para aprovação"
            )

        return {
            "success": True,
            "message": "Horário aprovado com sucesso",
            "updated_records": cursor.rowcount,
            "status": "approved"
        }

    except HTTPException:
        connection.rollback()
        raise

    finally:
        cursor.close()


@router.post("/reject-batch")
def reject_schedule_batch(
    request: ScheduleApprovalRequest,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director"]))
):
    if request.end_date < request.start_date:
        raise HTTPException(
            status_code=400,
            detail="A data final deve ser igual ou posterior à data inicial"
        )

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        cursor.execute("""
            UPDATE Tbl_GeneratedSchedule gs
            JOIN Tbl_SchoolCalendar sc
                ON gs.CalendarID = sc.CalendarID
            SET
                gs.Status = 'cancelled',
                gs.ChangeUsername = %s
            WHERE gs.ClassID = %s
              AND sc.CalendarDate BETWEEN %s AND %s
              AND gs.Status = 'draft'
        """, (
            audit_username,
            request.class_id,
            request.start_date,
            request.end_date
        ))

        connection.commit()

        if cursor.rowcount == 0:
            raise HTTPException(
                status_code=404,
                detail="Não foram encontrados horários em rascunho para rejeição"
            )

        return {
            "success": True,
            "message": "Horário rejeitado com sucesso",
            "updated_records": cursor.rowcount,
            "status": "cancelled",
            "reason": request.reason
        }

    except HTTPException:
        connection.rollback()
        raise

    finally:
        cursor.close()


@router.get("/export/excel")
def export_schedule_excel(
    class_id: Optional[int] = Query(default=None),
    professor_id: Optional[int] = Query(default=None),
    room_id: Optional[int] = Query(default=None),
    status: str = Query(default="approved"),
    start_date: Optional[date] = Query(default=None),
    end_date: Optional[date] = Query(default=None),
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        rows = fetch_schedule_export_rows(
            cursor=cursor,
            class_id=class_id,
            professor_id=professor_id,
            room_id=room_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Horário"

        title = "Academia360 - Exportação de Horário"
        sheet.merge_cells("A1:J1")
        title_cell = sheet["A1"]
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color="1C29E1")
        title_cell.alignment = Alignment(horizontal="center")

        generated_cell = sheet["A2"]
        generated_cell.value = f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        generated_cell.font = Font(italic=True, color="666666")

        headers = [
            "Turma",
            "Curso",
            "Ano letivo",
            "Data",
            "Início",
            "Fim",
            "Disciplina",
            "Professor",
            "Sala",
            "Estado",
        ]
        sheet.append([])
        sheet.append(headers)
        header_row = 4

        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1C29E1")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            sheet.append([
                row.get("class_name"),
                row.get("course_name"),
                row.get("school_year"),
                _format_export_value(row.get("schedule_date")),
                _format_export_value(row.get("start_time")),
                _format_export_value(row.get("end_time")),
                row.get("discipline_name"),
                row.get("professor_name"),
                row.get("room_name"),
                _status_label(row.get("status")),
            ])

        widths = [18, 28, 16, 14, 10, 10, 28, 28, 18, 14]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:J{sheet.max_row}"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = _schedule_export_filename(rows, "xlsx")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    finally:
        cursor.close()


@router.get("/export/pdf")
def export_schedule_pdf(
    class_id: Optional[int] = Query(default=None),
    professor_id: Optional[int] = Query(default=None),
    room_id: Optional[int] = Query(default=None),
    status: str = Query(default="approved"),
    start_date: Optional[date] = Query(default=None),
    end_date: Optional[date] = Query(default=None),
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        rows = fetch_schedule_export_rows(
            cursor=cursor,
            class_id=class_id,
            professor_id=professor_id,
            room_id=room_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
        )

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.0 * cm,
            leftMargin=1.0 * cm,
            topMargin=1.0 * cm,
            bottomMargin=1.0 * cm,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Academia360 - Horário", styles["Title"]),
            Paragraph(f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
            Spacer(1, 0.35 * cm),
        ]

        data = [["Data", "Início", "Fim", "Turma", "Disciplina", "Professor", "Sala", "Estado"]]
        for row in rows:
            data.append([
                _format_export_value(row.get("schedule_date")),
                _format_export_value(row.get("start_time")),
                _format_export_value(row.get("end_time")),
                row.get("class_name") or "",
                row.get("discipline_name") or "",
                row.get("professor_name") or "",
                row.get("room_name") or "",
                _status_label(row.get("status")),
            ])

        table = Table(data, repeatRows=1, colWidths=[2.2*cm, 1.5*cm, 1.5*cm, 3.0*cm, 4.0*cm, 4.2*cm, 2.8*cm, 2.2*cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C29E1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDE2FF")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FC")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))

        elements.append(table)
        document.build(elements)
        buffer.seek(0)

        filename = _schedule_export_filename(rows, "pdf")
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    finally:
        cursor.close()


@router.get("/{schedule_id}")
def get_schedule_record(
    schedule_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin", "director", "secretary", "professor"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                gs.ScheduleID AS id,
                gs.ClassID AS class_id,
                cl.Name AS class_name,
                c.CourseID AS course_id,
                c.Code AS course_code,
                c.Name AS course_name,
                sy.SchoolYearID AS school_year_id,
                sy.Name AS school_year,
                gs.DisciplineCourseYearID AS discipline_course_year_id,
                d.DisciplineID AS discipline_id,
                d.Name AS discipline_name,
                d.Code AS discipline_code,
                gs.ProfessorID AS professor_id,
                u.FullName AS professor_name,
                u.Email AS professor_email,
                gs.RoomID AS room_id,
                r.Name AS room_name,
                gs.CalendarID AS calendar_id,
                sc.CalendarDate AS schedule_date,
                gs.StartTime AS start_time,
                gs.EndTime AS end_time,
                gs.Status AS status,
                gs.InsertUsername AS insert_username,
                gs.InsertDate AS insert_date,
                gs.ChangeUsername AS change_username,
                gs.ChangeDate AS change_date
            FROM Tbl_GeneratedSchedule gs
            JOIN Tbl_Classes cl ON gs.ClassID = cl.ClassID
            JOIN Tbl_Courses c ON cl.CourseID = c.CourseID
            JOIN Tref_SchoolYears sy ON cl.SchoolYearID = sy.SchoolYearID
            JOIN trx_Discipline_CourseYear dc 
                ON gs.DisciplineCourseYearID = dc.DisciplineCourseYearID
            JOIN Tbl_Disciplines d 
                ON dc.DisciplineID = d.DisciplineID
            JOIN Tbl_Professors p 
                ON gs.ProfessorID = p.ProfessorID
            JOIN Tbl_Users u 
                ON p.UserID = u.UserID
            JOIN Tbl_Rooms r 
                ON gs.RoomID = r.RoomID
            JOIN Tbl_SchoolCalendar sc 
                ON gs.CalendarID = sc.CalendarID
            WHERE gs.ScheduleID = %s
        """, (schedule_id,))

        schedule = cursor.fetchone()

        if schedule is None:
            raise HTTPException(status_code=404, detail="Registo de horário não encontrado")

        return schedule

    finally:
        cursor.close()


@router.post("")
def create_schedule_record(
    schedule: ScheduleCreate,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin"]))
):
    validate_time_range(schedule.start_time, schedule.end_time)

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        validate_schedule_dependencies(
            cursor=cursor,
            class_id=schedule.class_id,
            discipline_course_year_id=schedule.discipline_course_year_id,
            professor_id=schedule.professor_id,
            room_id=schedule.room_id,
            calendar_id=schedule.calendar_id
        )

        validate_schedule_conflicts(
            cursor=cursor,
            class_id=schedule.class_id,
            professor_id=schedule.professor_id,
            room_id=schedule.room_id,
            calendar_id=schedule.calendar_id,
            start_time=schedule.start_time,
            end_time=schedule.end_time
        )

        cursor.execute("""
            INSERT INTO Tbl_GeneratedSchedule (
                ClassID,
                DisciplineCourseYearID,
                ProfessorID,
                RoomID,
                CalendarID,
                StartTime,
                EndTime,
                Status,
                InsertUsername
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            schedule.class_id,
            schedule.discipline_course_year_id,
            schedule.professor_id,
            schedule.room_id,
            schedule.calendar_id,
            schedule.start_time,
            schedule.end_time,
            schedule.status,
            audit_username
        ))

        connection.commit()

        return {
            "message": "Registo de horário criado com sucesso",
            "schedule_id": cursor.lastrowid
        }

    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))

    finally:
        cursor.close()


@router.put("/{schedule_id}")
def update_schedule_record(
    schedule_id: int,
    schedule: ScheduleUpdate,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin"]))
):
    data = model_to_dict(schedule)

    if not data:
        raise HTTPException(status_code=400, detail="Nenhum campo indicado para atualização")

    cursor = connection.cursor(dictionary=True)
    audit_username = get_audit_username(current_user)

    try:
        existing_schedule = get_existing_schedule(cursor, schedule_id)

        merged_schedule = {
            "class_id": data.get("class_id", existing_schedule["ClassID"]),
            "discipline_course_year_id": data.get(
                "discipline_course_year_id",
                existing_schedule["DisciplineCourseYearID"]
            ),
            "professor_id": data.get("professor_id", existing_schedule["ProfessorID"]),
            "room_id": data.get("room_id", existing_schedule["RoomID"]),
            "calendar_id": data.get("calendar_id", existing_schedule["CalendarID"]),
            "start_time": data.get("start_time", existing_schedule["StartTime"]),
            "end_time": data.get("end_time", existing_schedule["EndTime"])
        }

        validate_time_range(
            merged_schedule["start_time"],
            merged_schedule["end_time"]
        )

        validate_schedule_dependencies(
            cursor=cursor,
            class_id=merged_schedule["class_id"],
            discipline_course_year_id=merged_schedule["discipline_course_year_id"],
            professor_id=merged_schedule["professor_id"],
            room_id=merged_schedule["room_id"],
            calendar_id=merged_schedule["calendar_id"]
        )

        validate_schedule_conflicts(
            cursor=cursor,
            class_id=merged_schedule["class_id"],
            professor_id=merged_schedule["professor_id"],
            room_id=merged_schedule["room_id"],
            calendar_id=merged_schedule["calendar_id"],
            start_time=merged_schedule["start_time"],
            end_time=merged_schedule["end_time"],
            excluded_schedule_id=schedule_id
        )

        field_map = {
            "class_id": "ClassID",
            "discipline_course_year_id": "DisciplineCourseYearID",
            "professor_id": "ProfessorID",
            "room_id": "RoomID",
            "calendar_id": "CalendarID",
            "start_time": "StartTime",
            "end_time": "EndTime",
            "status": "Status"
        }

        set_clauses = []
        values = []

        for api_field, db_field in field_map.items():
            if api_field in data:
                set_clauses.append(f"{db_field} = %s")
                values.append(data[api_field])

        if not set_clauses:
            raise HTTPException(status_code=400, detail="Nenhum campo válido indicado para atualização")

        set_clauses.append("ChangeUsername = %s")
        values.append(audit_username)

        values.append(schedule_id)

        cursor.execute(f"""
            UPDATE Tbl_GeneratedSchedule
            SET {", ".join(set_clauses)}
            WHERE ScheduleID = %s
        """, tuple(values))

        connection.commit()

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Registo de horário não encontrado")

        return {
            "message": "Registo de horário atualizado com sucesso",
            "schedule_id": schedule_id
        }

    except IntegrityError as error:
        connection.rollback()
        raise HTTPException(status_code=400, detail=str(error))

    finally:
        cursor.close()


@router.delete("/{schedule_id}")
def delete_schedule_record(
    schedule_id: int,
    connection=Depends(get_db),
    current_user=Depends(require_roles(["admin"]))
):
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
            DELETE FROM Tbl_GeneratedSchedule
            WHERE ScheduleID = %s
        """, (schedule_id,))

        connection.commit()

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Registo de horário não encontrado")

        return {
            "message": "Registo de horário eliminado com sucesso",
            "schedule_id": schedule_id
        }

    except IntegrityError:
        connection.rollback()
        raise HTTPException(
            status_code=400,
            detail="O registo de horário não pode ser eliminado porque está a ser usado por outro registo"
        )

    finally:
        cursor.close()